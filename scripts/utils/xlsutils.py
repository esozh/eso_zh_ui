#!/usr/bin/env python
# -*- coding:utf-8 -*-
# File          : xlsutils.py
# Author        : bssthu
# Project       : eso_zh_ui
# Description   :
# 


import os
import shutil
import xlrd
import xlsxwriter
import openpyxl
from openpyxl.writer.write_only import WriteOnlyCell


def load_xls(file_path):
    """读取 Excel 文件中的数据。

    Args:
        file_path (str): xlsx 文件的路径

    Returns:
        data (list[list[str]]): 工作表中的内容
    """
    data = []
    with xlrd.open_workbook(file_path) as workbook:
        sheet = workbook.sheet_by_index(0)
        nrows = sheet.nrows
        ncols = sheet.ncols
        for curr_row in range(0, nrows):
            data.append([str(sheet.cell(curr_row, curr_col).value) for curr_col in range(0, ncols)])
    return data


def save_xlsx(file_path, data, header=None, col_id=None):
    """将输入保存到 Excel 文件中。使用文件模板

    全部保存为文本。

    Args:
        file_path (str): xlsx 文件的路径
        data (list[list]): 要保存的数据，二维
        header (list): 第一行
        col_id (list[int]): data 中列号到 xlsx 中列号的映射
    """
    if len(data) <= 0:
        return create_xlsx(file_path, data, header, col_id)

    cd = os.path.dirname(os.path.abspath(__file__))
    num_col = len(data[0])
    if num_col == 9:
        shutil.copy(os.path.join(cd, 'data/prototype_list.xlsx'), file_path)
    elif num_col == 12:
        shutil.copy(os.path.join(cd, 'data/prototype_pair.xlsx'), file_path)
    else:
        return create_xlsx(file_path, data, header, col_id)

    if col_id is None:
        col_id = list(range(0, num_col))
    max_col_id = max(col_id)

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name('sheet 1')

    # 格式
    fonts = [sheet.cell(row=1, column=i + 1).font.copy() for i in col_id]
    fills = [sheet.cell(row=1, column=i + 1).fill.copy() for i in col_id]
    alignments = [sheet.cell(row=1, column=i + 1).alignment.copy() for i in col_id]
    number_formats = [sheet.cell(row=1, column=i + 1).number_format for i in col_id]

    # 写入内容
    row_id = 1
    for row in data:
        row_cells = ['', ] * (max_col_id + 1)
        for j in range(0, num_col):
            cell = WriteOnlyCell(sheet, value=str(row[j]))
            cell.font = fonts[j]
            cell.fill = fills[j]
            cell.alignment = alignments[j]
            cell.number_format = number_formats[j]
            row_cells[col_id[j]] = cell
        sheet.append(row_cells)
        row_id += 1

    workbook.save(file_path)


def create_xlsx(file_path, data, header=None, col_id=None):
    """将输入保存到新建的 Excel 文件中。

    全部保存为文本。

    Args:
        file_path (str): xlsx 文件的路径
        data (list[list]): 要保存的数据，二维
        header (list): 第一行
        col_id (list[int]): data 中列号到 xlsx 中列号的映射
    """
    workbook = xlsxwriter.Workbook(file_path)
    sheet = workbook.add_worksheet('sheet 1')

    row_id = 0
    # 保存第一行
    if header is not None:
        for j in range(0, len(header)):
            sheet.write(row_id, j, str(header[j]))
        row_id += 1

    # 检查有没有数据
    if len(data) <= 0:
        workbook.close()
        return

    num_col = len(data[0])
    if col_id is None:
        col_id = list(range(0, num_col))
    # 保存其他行
    for row in data:
        for j in range(0, num_col):
            sheet.write(row_id, col_id[j], str(row[j]))
        row_id += 1

    workbook.close()
